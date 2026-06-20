#!/usr/bin/env python3
"""
Compare new NMA list (/Users/debo/Downloads/NMA Full List and Invalid List.xlsx)
against existing scripts/doctors-import.csv and flag duplicates.

Match strategies (priority):
  1. Exact email match (case-insensitive, trimmed)  -> STRONG
  2. Normalized phone match (last 10 digits)         -> STRONG
  3. Normalized full-name match                       -> MEDIUM
  4. Fuzzy name (token-set, >=85)                    -> WEAK (flag for human review)

Outputs three CSVs into /tmp:
  - nma_duplicates_strong.csv     (email or phone match)
  - nma_duplicates_name.csv       (exact/fuzzy name only, no email/phone match)
  - nma_unique_new.csv            (no match at all - safe to add)
Also prints summary counts.
"""
import csv
import os
import re
import sys
from difflib import SequenceMatcher

import openpyxl

EXISTING = "/Users/debo/consult-for-africa/scripts/doctors-import.csv"
NEW = "/Users/debo/Downloads/NMA Full List and Invalid List.xlsx"
OUT_STRONG = "/tmp/nma_duplicates_strong.csv"
OUT_NAME = "/tmp/nma_duplicates_name.csv"
OUT_UNIQUE = "/tmp/nma_unique_new.csv"


def norm_email(e):
    if not e:
        return ""
    return str(e).strip().lower().replace("\xa0", "").replace(" ", "")


def norm_phone(p):
    if not p:
        return ""
    digits = re.sub(r"\D", "", str(p))
    # Strip country code 234 if present
    if digits.startswith("234"):
        digits = digits[3:]
    # Strip leading 0
    if digits.startswith("0"):
        digits = digits[1:]
    # Require at least 9 digits to count as a real phone match (avoids "803", "23512" noise)
    if len(digits) < 9:
        return ""
    return digits[-10:]


def norm_name(n):
    if not n:
        return ""
    s = str(n).strip().lower()
    # strip Dr / DR / Doctor prefix
    s = re.sub(r"^(dr\.?|doctor)\s+", "", s)
    s = re.sub(r"^dr\s+dr\.?\s+", "", s)  # "Dr DR ANTHONY..."
    s = re.sub(r"^(dr\.?|doctor)\s+", "", s)
    # remove punctuation
    s = re.sub(r"[^\w\s]", " ", s)
    # collapse whitespace
    s = re.sub(r"\s+", " ", s).strip()
    return s


def name_tokens(n):
    return set(norm_name(n).split())


def fuzzy_ratio(a, b):
    return SequenceMatcher(None, a, b).ratio()


# --- Load existing ---
existing_rows = []
with open(EXISTING, newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for r in reader:
        existing_rows.append(r)

print(f"Loaded existing: {len(existing_rows)} rows")

email_idx = {}
phone_idx = {}
name_idx = {}  # normalized name -> list of rows
name_token_idx = []  # list of (token_set, name, row) for fuzzy

for r in existing_rows:
    e = norm_email(r.get("Email"))
    p = norm_phone(r.get("Phone"))
    n = norm_name(r.get("Full Name"))
    if e:
        email_idx.setdefault(e, []).append(r)
    if p:
        phone_idx.setdefault(p, []).append(r)
    if n:
        name_idx.setdefault(n, []).append(r)
        name_token_idx.append((name_tokens(r.get("Full Name")), n, r))

print(f"  email keys: {len(email_idx)}, phone keys: {len(phone_idx)}, name keys: {len(name_idx)}")

# --- Load new ---
wb = openpyxl.load_workbook(NEW, data_only=True)
new_rows = []
for sn in wb.sheetnames:
    ws = wb[sn]
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        if all(c is None for c in row):
            continue
        rec = dict(zip(headers, row))
        rec["__sheet"] = sn
        new_rows.append(rec)

print(f"Loaded new: {len(new_rows)} rows across {len(wb.sheetnames)} sheet(s)")

# --- Match ---
strong_dupes = []
name_dupes = []
unique_new = []

for r in new_rows:
    name = r.get("Full Name") or ""
    email = r.get("Email ") or r.get("Email") or ""
    phone = r.get(" Phone") or r.get("Phone") or ""
    pow_ = r.get("Place of Work") or ""
    spec = r.get("Specialty") or ""
    sheet = r.get("__sheet", "")

    e = norm_email(email)
    p = norm_phone(phone)
    n = norm_name(name)

    matches = []
    match_type = []

    if e and e in email_idx:
        matches.extend(email_idx[e])
        match_type.append("email")
    if p and p in phone_idx:
        matches.extend(phone_idx[p])
        match_type.append("phone")

    if matches:
        existing = matches[0]
        strong_dupes.append({
            "match_type": "+".join(match_type),
            "new_sheet": sheet,
            "new_name": name,
            "new_email": email,
            "new_phone": phone,
            "new_place": pow_,
            "new_specialty": spec,
            "existing_sno": existing.get("S/No"),
            "existing_name": existing.get("Full Name"),
            "existing_email": existing.get("Email"),
            "existing_phone": existing.get("Phone"),
            "existing_specialty": existing.get("Specialty"),
        })
        continue

    # Exact name match (no email/phone match)
    if n and n in name_idx:
        existing = name_idx[n][0]
        name_dupes.append({
            "match_type": "exact_name",
            "score": 1.0,
            "new_sheet": sheet,
            "new_name": name,
            "new_email": email,
            "new_phone": phone,
            "new_place": pow_,
            "new_specialty": spec,
            "existing_sno": existing.get("S/No"),
            "existing_name": existing.get("Full Name"),
            "existing_email": existing.get("Email"),
            "existing_phone": existing.get("Phone"),
            "existing_specialty": existing.get("Specialty"),
        })
        continue

    # Fuzzy name (token-set jaccard + sequence)
    if n:
        toks = name_tokens(name)
        best = None
        best_score = 0.0
        for etoks, en, er in name_token_idx:
            if not etoks or not toks:
                continue
            # Quick filter: require at least 2 shared tokens of length>=3
            shared = {t for t in (toks & etoks) if len(t) >= 3}
            if len(shared) < 2:
                continue
            # Jaccard
            jac = len(toks & etoks) / len(toks | etoks)
            if jac < 0.6:
                continue
            # Sequence ratio on full string for confirmation
            seq = fuzzy_ratio(n, en)
            score = 0.5 * jac + 0.5 * seq
            if score > best_score:
                best_score = score
                best = er
        if best and best_score >= 0.85:
            name_dupes.append({
                "match_type": "fuzzy_name",
                "score": round(best_score, 3),
                "new_sheet": sheet,
                "new_name": name,
                "new_email": email,
                "new_phone": phone,
                "new_place": pow_,
                "new_specialty": spec,
                "existing_sno": best.get("S/No"),
                "existing_name": best.get("Full Name"),
                "existing_email": best.get("Email"),
                "existing_phone": best.get("Phone"),
                "existing_specialty": best.get("Specialty"),
            })
            continue

    unique_new.append({
        "sheet": sheet,
        "name": name,
        "email": email,
        "phone": phone,
        "place": pow_,
        "specialty": spec,
    })


def write_csv(path, rows, fields):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow(r)


write_csv(
    OUT_STRONG,
    strong_dupes,
    ["match_type", "new_sheet", "new_name", "new_email", "new_phone", "new_place", "new_specialty",
     "existing_sno", "existing_name", "existing_email", "existing_phone", "existing_specialty"],
)
write_csv(
    OUT_NAME,
    name_dupes,
    ["match_type", "score", "new_sheet", "new_name", "new_email", "new_phone", "new_place", "new_specialty",
     "existing_sno", "existing_name", "existing_email", "existing_phone", "existing_specialty"],
)
write_csv(
    OUT_UNIQUE,
    unique_new,
    ["sheet", "name", "email", "phone", "place", "specialty"],
)

print()
print(f"STRONG dupes (email or phone match): {len(strong_dupes)}  -> {OUT_STRONG}")
print(f"NAME dupes (exact or fuzzy name only): {len(name_dupes)}  -> {OUT_NAME}")
print(f"UNIQUE new (safe to add):              {len(unique_new)}  -> {OUT_UNIQUE}")

# Sheet breakdown
from collections import Counter
print()
print("New rows by sheet:")
for sn, c in Counter(r.get("__sheet") for r in new_rows).items():
    print(f"  {sn}: {c}")
print("Strong dupes by sheet:")
for sn, c in Counter(r["new_sheet"] for r in strong_dupes).items():
    print(f"  {sn}: {c}")
print("Name dupes by sheet:")
for sn, c in Counter(r["new_sheet"] for r in name_dupes).items():
    print(f"  {sn}: {c}")
print("Unique by sheet:")
for sn, c in Counter(r["sheet"] for r in unique_new).items():
    print(f"  {sn}: {c}")
